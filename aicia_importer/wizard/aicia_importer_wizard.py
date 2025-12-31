import logging
from base64 import b64decode
from io import BytesIO

from odoo import _, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.warning("openpyxl library not installed. Please install it to use this module.")
    openpyxl = None


class AiciaImporterWizard(models.TransientModel):
    _name = "aicia.importer.wizard"
    _description = "Importador de Proveedores y Clientes AICIA"

    # Mapeo de códigos de provincia
    PROVINCE_MAP = {
        '1': 'ALAVA',
        '2': 'ALBACETE',
        '3': 'ALACANT/ALICANTE',
        '4': 'ALMERIA',
        '5': 'AVILA',
        '6': 'BADAJOZ',
        '7': 'ILLES BALEARS',
        '8': 'BARCELONA',
        '9': 'BURGOS',
        '10': 'CACERES',
        '11': 'CADIZ',
        '12': 'CASTELLON DE LA PLANA',
        '13': 'CIUDAD REAL',
        '14': 'CORDOBA',
        '15': 'A CORUÑA',
        '16': 'CUENCA',
        '17': 'GIRONA',
        '18': 'GRANADA',
        '19': 'GUADALAJARA',
        '20': 'GUIPUZCOA',
        '21': 'HUELVA',
        '22': 'HUESCA',
        '23': 'JAEN',
        '24': 'LEON',
        '25': 'LLEIDA',
        '26': 'LA RIOJA',
        '27': 'LUGO',
        '28': 'MADRID',
        '29': 'MALAGA',
        '30': 'MURCIA',
        '31': 'NAFARROA/NAVARRA',
        '32': 'OURENSE',
        '33': 'ASTURIAS',
        '34': 'PALENCIA',
        '35': 'LAS PALMAS',
        '36': 'PONTEVEDRA',
        '37': 'SALAMANCA',
        '38': 'STA. CRUZ TENERIFE',
        '39': 'CANTABRIA',
        '40': 'SEGOVIA',
        '41': 'SEVILLA',
        '42': 'SORIA',
        '43': 'TARRAGONA',
        '44': 'TERUEL',
        '45': 'TOLEDO',
        '46': 'VALENCIA',
        '47': 'VALLADOLID',
        '48': 'VIZCAYA',
        '49': 'ZAMORA',
        '50': 'ZARAGOZA',
        '51': 'CEUTA',
        '52': 'MELILLA',
    }

    data_file_suppliers = fields.Binary(
        string="Archivo de Proveedores",
        help="Seleccione el archivo Excel con los proveedores a importar.",
    )
    filename_suppliers = fields.Char(string="Nombre del archivo de proveedores")

    data_file_customers = fields.Binary(
        string="Archivo de Clientes",
        help="Seleccione el archivo Excel con los clientes a importar.",
    )
    filename_customers = fields.Char(string="Nombre del archivo de clientes")

    data_file_employees = fields.Binary(
        string="Archivo de Personal",
        help="Seleccione el archivo Excel con el personal a importar.",
    )
    filename_employees = fields.Char(string="Nombre del archivo de personal")

    update_existing = fields.Boolean(
        string="Actualizar existentes",
        default=False,
        help="Si está marcado, actualiza los contactos existentes. Si no, solo crea nuevos.",
    )

    def import_suppliers(self):
        """Procesa los archivos Excel e importa proveedores y/o clientes."""
        self.ensure_one()

        if not openpyxl:
            raise UserError(
                _("La librería 'openpyxl' no está instalada. "
                  "Por favor, instálela con: pip install openpyxl")
            )

        if not self.data_file_suppliers and not self.data_file_customers and not self.data_file_employees:
            raise UserError(_("Por favor, seleccione al menos un archivo para importar."))

        _logger.info("=" * 80)
        _logger.info("INICIO DE IMPORTACIÓN AICIA")
        _logger.info("=" * 80)

        # Crear mapeo de provincias desde el diccionario interno
        _logger.info(">>> Cargando mapeo de provincias")
        province_map = self._get_province_mapping()
        _logger.info(f">>> PROVINCIAS: {len(province_map)} códigos mapeados")

        total_created = 0
        total_updated = 0
        total_errors = 0
        all_errors = []

        # Importar proveedores si se proporciona el archivo
        if self.data_file_suppliers:
            _logger.info(">>> Iniciando importación de PROVEEDORES")
            result = self._import_file(self.data_file_suppliers, is_supplier=True, province_map=province_map)
            total_created += result['created']
            total_updated += result['updated']
            total_errors += result['errors']
            all_errors.extend([f"[Proveedores] {e}" for e in result['error_list']])
            _logger.info(f">>> PROVEEDORES: {result['created']} creados, {result['updated']} actualizados, {result['errors']} errores")

        # Importar clientes si se proporciona el archivo
        if self.data_file_customers:
            _logger.info(">>> Iniciando importación de CLIENTES")
            result = self._import_file(self.data_file_customers, is_supplier=False, province_map=province_map)
            total_created += result['created']
            total_updated += result['updated']
            total_errors += result['errors']
            all_errors.extend([f"[Clientes] {e}" for e in result['error_list']])
            _logger.info(f">>> CLIENTES: {result['created']} creados, {result['updated']} actualizados, {result['errors']} errores")

        # Importar personal si se proporciona el archivo
        if self.data_file_employees:
            _logger.info(">>> Iniciando importación de PERSONAL")
            result = self._import_employees(self.data_file_employees, province_map=province_map)
            total_created += result['created']
            total_updated += result['updated']
            total_errors += result['errors']
            all_errors.extend([f"[Personal] {e}" for e in result['error_list']])
            _logger.info(f">>> PERSONAL: {result['created']} creados, {result['updated']} actualizados, {result['errors']} errores")

        _logger.info("=" * 80)
        _logger.info(f"FIN DE IMPORTACIÓN - Total: {total_created} creados, {total_updated} actualizados, {total_errors} errores")
        _logger.info("=" * 80)

        # Mensaje de resultado
        message = _(
            "Importación completada:\n"
            "- Contactos creados: %s\n"
            "- Contactos actualizados: %s\n"
            "- Errores: %s"
        ) % (total_created, total_updated, total_errors)

        if all_errors:
            message += "\n\nDetalles de errores:\n" + "\n".join(all_errors[:10])
            if len(all_errors) > 10:
                message += f"\n... y {len(all_errors) - 10} errores más."

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Importación de Proveedores y Clientes'),
                'message': message,
                'type': 'success' if total_errors == 0 else 'warning',
                'sticky': True,
            }
        }

    def _get_province_mapping(self):
        """Crea un mapeo código -> res.country.state usando el diccionario PROVINCE_MAP."""
        province_map = {}
        country_es = self.env['res.country'].search([('code', '=', 'ES')], limit=1)

        if not country_es:
            _logger.warning("No se encontró el país España (ES) en la base de datos")
            return {}

        # Mapeo específico de códigos a códigos de provincia en Odoo
        code_mapping = {
            '1': 'VI',  # ALAVA -> Araba/Álava
            '2': 'AB',  # ALBACETE
            '3': 'A',  # ALACANT/ALICANTE -> Alacant (Alicante)
            '4': 'AL',  # ALMERIA -> Almería
            '5': 'AV',  # AVILA -> Ávila
            '6': 'BA',  # BADAJOZ -> Badajoz
            '7': 'PM',  # ILLES BALEARS -> Illes Balears
            '8': 'B',  # BARCELONA
            '9': 'BU',  # BURGOS
            '01': 'VI',      # ALAVA -> Araba/Álava
            '02': 'AB',      # ALBACETE
            '03': 'A',       # ALACANT/ALICANTE -> Alacant (Alicante)
            '04': 'AL',      # ALMERIA -> Almería
            '05': 'AV',      # AVILA -> Ávila
            '06': 'BA',      # BADAJOZ -> Badajoz
            '07': 'PM',      # ILLES BALEARS -> Illes Balears
            '08': 'B',       # BARCELONA
            '09': 'BU',      # BURGOS
            '10': 'CC',     # CACERES -> Cáceres
            '11': 'CA',     # CADIZ -> Cádiz
            '12': 'CS',     # CASTELLON -> Castelló
            '13': 'CR',     # CIUDAD REAL
            '14': 'CO',     # CORDOBA -> Córdoba
            '15': 'C',      # A CORUÑA
            '16': 'CU',     # CUENCA
            '17': 'GI',     # GIRONA
            '18': 'GR',     # GRANADA
            '19': 'GU',     # GUADALAJARA
            '20': 'SS',     # GUIPUZCOA -> Gipuzkoa
            '21': 'H',      # HUELVA
            '22': 'HU',     # HUESCA
            '23': 'J',      # JAEN -> Jaén
            '24': 'LE',     # LEON -> León
            '25': 'L',      # LLEIDA
            '26': 'LO',     # LA RIOJA
            '27': 'LU',     # LUGO
            '28': 'M',      # MADRID
            '29': 'MA',     # MALAGA -> Málaga
            '30': 'MU',     # MURCIA
            '31': 'NA',     # NAFARROA/NAVARRA -> Navarra
            '32': 'OR',     # OURENSE
            '33': 'O',      # ASTURIAS
            '34': 'P',      # PALENCIA
            '35': 'GC',     # LAS PALMAS
            '36': 'PO',     # PONTEVEDRA
            '37': 'SA',     # SALAMANCA
            '38': 'TF',     # STA. CRUZ TENERIFE -> Santa Cruz de Tenerife
            '39': 'S',      # CANTABRIA
            '40': 'SG',     # SEGOVIA
            '41': 'SE',     # SEVILLA
            '42': 'SO',     # SORIA
            '43': 'T',      # TARRAGONA
            '44': 'TE',     # TERUEL
            '45': 'TO',     # TOLEDO
            '46': 'V',      # VALENCIA -> València
            '47': 'VA',     # VALLADOLID
            '48': 'BI',     # VIZCAYA -> Bizkaia
            '49': 'ZA',     # ZAMORA
            '50': 'Z',      # ZARAGOZA
            '51': 'CE',     # CEUTA
            '52': 'ME',     # MELILLA
        }

        for codigo, odoo_code in code_mapping.items():
            # Buscar la provincia en Odoo por código
            state = self.env['res.country.state'].search([
                ('country_id', '=', country_es.id),
                ('code', '=', odoo_code)
            ], limit=1)

            if state:
                province_map[codigo] = state.id
                _logger.debug(f"Provincia mapeada: código {codigo} ({self.PROVINCE_MAP.get(codigo)}) -> {state.name} (código Odoo: {odoo_code}, ID: {state.id})")
            else:
                _logger.warning(f"No se encontró provincia en Odoo para código: {codigo}, código Odoo: {odoo_code}")

        _logger.info(f"Mapeo de provincias completado: {len(province_map)}/52 provincias mapeadas")
        return province_map


    def _import_file(self, data_file, is_supplier=True, province_map=None):
        """Importa un archivo Excel de proveedores o clientes."""
        try:
            # Decodificar el archivo
            file_data = b64decode(data_file)
            workbook = openpyxl.load_workbook(BytesIO(file_data))
            sheet = workbook.active

            # Obtener las filas del archivo
            rows = list(sheet.iter_rows(values_only=True))

            if len(rows) < 2:
                raise UserError(_("El archivo no contiene datos para importar."))

            # La primera fila son los encabezados
            headers = rows[0]
            data_rows = rows[1:]
            total_rows = len(data_rows)

            tipo = "proveedores" if is_supplier else "clientes"
            _logger.info(f"Iniciando importación de {total_rows} {tipo}...")

            # Procesar cada fila
            created_count = 0
            updated_count = 0
            error_count = 0
            errors = []
            batch_size = 50  # Hacer commit cada 50 registros

            for idx, row in enumerate(data_rows, start=2):  # start=2 porque la fila 1 es encabezado
                try:
                    # Log de progreso cada 100 registros
                    current = idx - 1
                    if current % 100 == 0:
                        _logger.info(f"Procesando {tipo}: {current}/{total_rows} ({int(current*100/total_rows)}%)")

                    result = self._process_partner_row(row, headers, is_supplier, province_map)
                    if result == 'created':
                        created_count += 1
                    elif result == 'updated':
                        updated_count += 1

                    # Hacer commit cada batch_size para que se vea el progreso
                    if current % batch_size == 0:
                        self.env.cr.commit()

                except Exception as e:
                    error_count += 1
                    error_msg = f"Fila {idx}: {str(e)}"
                    errors.append(error_msg)
                    _logger.error(error_msg)

            # Commit final
            self.env.cr.commit()
            _logger.info(f"Importación de {tipo} completada: {created_count} creados, {updated_count} actualizados, {error_count} errores")

            return {
                'created': created_count,
                'updated': updated_count,
                'errors': error_count,
                'error_list': errors,
            }

        except Exception as e:
            raise UserError(
                _("Error al procesar el archivo: %s") % str(e)
            )

    def _process_partner_row(self, row, headers, is_supplier=True, province_map=None):
        """Procesa una fila del Excel y crea o actualiza el proveedor o cliente."""
        # Mapear los datos de la fila
        data = dict(zip(headers, row))

        # Campos principales - Detectar si es proveedor o cliente por los campos
        if is_supplier:
            ref = str(data.get('ID_Proveedor', '')).strip() if data.get('ID_Proveedor') else False
        else:
            ref = str(data.get('ID_Cliente', '')).strip() if data.get('ID_Cliente') else False

        name = data.get('Nombre', '').strip() if data.get('Nombre') else False

        if not name:
            raise UserError(_("El campo 'Nombre' es obligatorio."))

        # Construir nombre completo si hay nombre2
        nombre2 = data.get('Nombre2', '').strip() if data.get('Nombre2') else ''
        if nombre2:
            full_name = f"{name} {nombre2}"
        else:
            full_name = name

        # Preparar valores para el contacto
        partner_vals = {
            'name': full_name,
            'is_company': True,
        }

        # Marcar como proveedor o cliente
        if is_supplier:
            partner_vals['supplier_rank'] = 1
        else:
            partner_vals['customer_rank'] = 1

        # Referencia interna
        if ref:
            partner_vals['ref'] = ref

        # NIF/CIF
        vat = data.get('CIF', '').strip() if data.get('CIF') else False
        if vat:
            # Si no empieza con el código del país, añadir ES
            if not vat[:2].isalpha():
                vat = 'ES' + vat
            partner_vals['vat'] = vat

        # Teléfonos - Concatenar teléfono fijo y móvil si ambos existen
        phone = data.get('Telefono', '').strip() if data.get('Telefono') else False
        mobile = data.get('TelefonoMovil', '').strip() if data.get('TelefonoMovil') else False

        if phone and mobile:
            # Si hay ambos, concatenar con separador
            partner_vals['phone'] = f"{phone} / {mobile}"
        elif phone:
            partner_vals['phone'] = phone
        elif mobile:
            partner_vals['phone'] = mobile

        # Email
        email = data.get('Correo_Electronico', '').strip() if data.get('Correo_Electronico') else False
        if email:
            partner_vals['email'] = email

        # Dirección
        street = data.get('Direccion', '').strip() if data.get('Direccion') else False
        street2 = data.get('Direccion2', '').strip() if data.get('Direccion2') else False

        if street:
            partner_vals['street'] = street
        if street2:
            partner_vals['street2'] = street2

        # Ciudad y código postal
        city = data.get('Localidad', '').strip() if data.get('Localidad') else False
        if city:
            partner_vals['city'] = city

        zip_code = data.get('Cod_Postal', '').strip() if data.get('Cod_Postal') else False
        if zip_code:
            partner_vals['zip'] = str(zip_code)

        # País - Por defecto España si no se especifica
        country = self.env['res.country'].search([('code', '=', 'ES')], limit=1)
        if country:
            partner_vals['country_id'] = country.id

        # Provincia - extraer del código postal (2 primeros dígitos)
        if province_map and zip_code:
            # Extraer los dos primeros caracteres del código postal
            try:
                zip_str = str(zip_code).strip()
                if len(zip_str) >= 2:
                    provincia_codigo = zip_str[:2]
                    # Eliminar ceros a la izquierda para códigos como "01", "02", etc.
                    provincia_codigo = str(int(provincia_codigo))

                    if provincia_codigo in province_map:
                        partner_vals['state_id'] = province_map[provincia_codigo]
                        _logger.debug(f"✓ Provincia asignada desde CP {zip_code}: código {provincia_codigo} -> state_id {province_map[provincia_codigo]}")
                    else:
                        _logger.warning(f"✗ Código de provincia no encontrado: '{provincia_codigo}' extraído del CP: {zip_code}")
                else:
                    _logger.warning(f"✗ Código postal demasiado corto para extraer provincia: '{zip_code}'")
            except Exception as e:
                _logger.warning(f"✗ Error al extraer provincia del código postal '{zip_code}': {str(e)}")

        # Comentarios/Observaciones
        comment = data.get('Observaciones', '').strip() if data.get('Observaciones') else ''
        comment2 = data.get('Observaciones2', '').strip() if data.get('Observaciones2') else ''

        if comment or comment2:
            full_comment = '\n'.join(filter(None, [comment, comment2]))
            partner_vals['comment'] = full_comment

        # Persona de contacto
        contact_person = data.get('Persona_Contacto', '').strip() if data.get('Persona_Contacto') else False

        # Buscar si el proveedor ya existe
        domain = []
        if ref:
            domain.append(('ref', '=', ref))
        elif vat:
            domain.append(('vat', '=', vat))
        else:
            domain.append(('name', '=', full_name))

        existing_partner = self.env['res.partner'].search(domain, limit=1)

        if existing_partner:
            if self.update_existing:
                existing_partner.write(partner_vals)

                # Si hay persona de contacto y no existe, crearla
                if contact_person:
                    self._create_contact_person(existing_partner, contact_person, data)

                # Crear o actualizar cuenta bancaria
                self._create_or_update_bank_account(existing_partner, data)

                return 'updated'
            else:
                return 'skipped'
        else:
            # Crear nuevo proveedor
            new_partner = self.env['res.partner'].create(partner_vals)

            # Si hay persona de contacto, crearla como contacto del proveedor
            if contact_person:
                self._create_contact_person(new_partner, contact_person, data)

            # Crear cuenta bancaria si hay datos
            self._create_or_update_bank_account(new_partner, data)

            return 'created'

    def _create_contact_person(self, partner, contact_name, data):
        """Crea una persona de contacto para el proveedor."""
        # Verificar si ya existe un contacto con ese nombre
        existing_contact = self.env['res.partner'].search([
            ('parent_id', '=', partner.id),
            ('name', '=', contact_name),
        ], limit=1)

        if not existing_contact:
            contact_vals = {
                'name': contact_name,
                'parent_id': partner.id,
                'type': 'contact',
            }

            # Si hay email específico para el contacto (aunque en este caso va al principal)
            email = data.get('Correo_Electronico', '').strip() if data.get('Correo_Electronico') else False
            if email:
                contact_vals['email'] = email

            self.env['res.partner'].create(contact_vals)

    def _create_or_update_bank_account(self, partner, data):
        """Crea o actualiza la cuenta bancaria del contacto."""
        try:
            # Obtener datos bancarios del Excel - convertir a string y limpiar
            id_banco = str(data.get('ID_Banco', '') or '').strip()
            id_sucursal = str(data.get('ID_Sucursal', '') or '').strip()
            digito_control = str(data.get('Digito_Control', '') or '').strip()
            numero_cuenta = str(data.get('Numero_Cuenta', '') or '').strip()
            iban_pais = str(data.get('IBAN_Pais', '') or 'ES').strip()
            iban_dc = str(data.get('IBAN_DC', '') or '').strip()

            _logger.debug(f"Datos bancarios para {partner.name}: Banco={id_banco}, Sucursal={id_sucursal}, DC={digito_control}, Cuenta={numero_cuenta}, IBAN_DC={iban_dc}")

            # Si no hay datos bancarios significativos, no crear cuenta
            if not numero_cuenta or numero_cuenta == '0' or numero_cuenta == '0000000000':
                _logger.debug(f"Sin datos bancarios válidos para {partner.name}")
                return

            # Construir número de cuenta bancaria
            # Formato antiguo español: BANCO-SUCURSAL-DC-CUENTA
            acc_number_parts = []
            if id_banco and id_banco != '0':
                acc_number_parts.append(id_banco.zfill(4))
            if id_sucursal and id_sucursal != '0':
                acc_number_parts.append(id_sucursal.zfill(4))
            if digito_control and digito_control != '0':
                acc_number_parts.append(digito_control.zfill(2))
            if numero_cuenta and numero_cuenta != '0':
                acc_number_parts.append(numero_cuenta.zfill(10))

            # Construir IBAN si tenemos los datos necesarios
            acc_number = None
            if len(acc_number_parts) == 4:
                # Formato CCC (Código Cuenta Cliente) español
                ccc = ''.join(acc_number_parts)
                # Construir IBAN: ES + DC + CCC
                if iban_dc and iban_dc != '0':
                    acc_number = f"{iban_pais}{iban_dc.zfill(2)}{ccc}"
                    _logger.debug(f"IBAN construido para {partner.name}: {acc_number}")
                else:
                    # Si no tenemos dígito de control del IBAN, usar el CCC directamente
                    acc_number = ccc
                    _logger.debug(f"CCC construido para {partner.name}: {acc_number}")
            elif numero_cuenta and numero_cuenta != '0':
                # Si solo tenemos número de cuenta, usar ese
                acc_number = numero_cuenta
                _logger.debug(f"Solo número cuenta para {partner.name}: {acc_number}")

            if not acc_number or acc_number == '0':
                _logger.debug(f"Número de cuenta vacío o cero para {partner.name}")
                return

            # Buscar si ya existe una cuenta bancaria para este partner
            existing_bank = self.env['res.partner.bank'].search([
                ('partner_id', '=', partner.id),
                ('acc_number', '=', acc_number),
            ], limit=1)

            if existing_bank:
                _logger.debug(f"Cuenta bancaria ya existe para {partner.name}: {acc_number}")
                # Actualizar para asegurar que allow_out_payment está activo
                if not existing_bank.allow_out_payment:
                    existing_bank.write({'allow_out_payment': True})
                    _logger.info(f"✓ Campo allow_out_payment activado para cuenta de {partner.name}")
                return
            else:
                # Crear nueva cuenta bancaria
                bank_vals = {
                    'partner_id': partner.id,
                    'acc_number': acc_number,
                    'allow_out_payment': True,  # Permitir pagos salientes
                }

                # Si es un IBAN válido (empieza con código país de 2 letras)
                if acc_number and len(acc_number) > 2 and acc_number[:2].isalpha():
                    bank_vals['acc_type'] = 'iban'

                self.env['res.partner.bank'].create(bank_vals)
                _logger.info(f"✓ Cuenta bancaria creada para {partner.name}: {acc_number} (allow_out_payment=True)")

        except Exception as e:
            # Si hay error creando la cuenta, registrar pero continuar
            _logger.warning(f"✗ Error creando cuenta bancaria para {partner.name}: {str(e)}")

    def _import_employees(self, file_data, province_map):
        """Importa empleados desde un archivo Excel."""
        result = {
            'created': 0,
            'updated': 0,
            'errors': 0,
            'error_list': []
        }

        try:
            # Decodificar y cargar el archivo Excel
            file_content = b64decode(file_data)
            workbook = openpyxl.load_workbook(BytesIO(file_content))
            sheet = workbook.active

            # Obtener los encabezados de las columnas
            headers = {}
            for col_idx, cell in enumerate(sheet[1], start=1):
                if cell.value:
                    headers[cell.value.strip()] = col_idx

            _logger.info(f">>> Columnas encontradas: {list(headers.keys())}")

            # Procesar las filas de datos (desde la fila 2)
            total_rows = sheet.max_row - 1
            _logger.info(f">>> Total de filas a procesar: {total_rows}")

            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Mostrar progreso cada 10 registros
                    if (row_idx - 1) % 10 == 0:
                        _logger.info(f">>> Procesando empleado {row_idx - 1}/{total_rows}")

                    # Mapear los datos de la fila
                    data = dict(zip(headers.keys(), row))

                    # Obtener valores de las celdas
                    codigo = str(data.get('Código Personal', '')).strip() if data.get('Código Personal') else ''
                    nombre = str(data.get('Nombre', '')).strip() if data.get('Nombre') else ''
                    apellido1 = str(data.get('Apellido1', '')).strip() if data.get('Apellido1') else ''
                    apellido2 = str(data.get('Apellido2', '')).strip() if data.get('Apellido2') else ''
                    nif = str(data.get('NIF', '')).strip() if data.get('NIF') else ''
                    direccion = str(data.get('Direccion', '')).strip() if data.get('Direccion') else ''
                    cod_postal = str(data.get('Cod_Postal', '')).strip() if data.get('Cod_Postal') else ''
                    poblacion = str(data.get('Población', '')).strip() if data.get('Población') else ''
                    telefono = str(data.get('Teléfono', '')).strip() if data.get('Teléfono') else ''
                    movil = str(data.get('Movil', '')).strip() if data.get('Movil') else ''
                    email = str(data.get('Email', '')).strip() if data.get('Email') else ''
                    nss = str(data.get('NSS', '')).strip() if data.get('NSS') else ''
                    fecha_antiguedad = data.get('Fecha Antiguedad', '')
                    cuenta_bancaria = str(data.get('Cuenta Bancaria', '')).strip() if data.get('Cuenta Bancaria') else ''
                    tipo_contrato = str(data.get('Tipo de contrato', '')).strip() if data.get('Tipo de contrato') else ''

                    # Validaciones básicas
                    if not nombre and not apellido1:
                        _logger.warning(f"Fila {row_idx}: Sin nombre ni apellido, saltando...")
                        continue

                    # Construir nombre completo
                    full_name_parts = []
                    if nombre:
                        full_name_parts.append(nombre)
                    if apellido1:
                        full_name_parts.append(apellido1)
                    if apellido2:
                        full_name_parts.append(apellido2)
                    full_name = ' '.join(full_name_parts)

                    # Buscar empleado existente por NIF o nombre
                    employee = None
                    if nif:
                        employee = self.env['hr.employee'].search([
                            ('identification_id', '=', nif)
                        ], limit=1)

                    if not employee and codigo:
                        # Buscar por código en campo work_contact_id.ref
                        partner = self.env['res.partner'].search([
                            ('ref', '=', codigo)
                        ], limit=1)
                        if partner:
                            employee = self.env['hr.employee'].search([
                                ('work_contact_id', '=', partner.id)
                            ], limit=1)

                    if not employee:
                        # Buscar por nombre completo
                        employee = self.env['hr.employee'].search([
                            ('name', '=', full_name)
                        ], limit=1)

                    # Preparar valores para el contacto de trabajo (res.partner)
                    partner_vals = {
                        'name': full_name,
                        'ref': codigo,
                        'vat': nif if nif else False,
                        'street': direccion if direccion else False,
                        'zip': cod_postal if cod_postal else False,
                        'city': poblacion if poblacion else False,
                        'email': email if email else False,
                        'country_id': self.env.ref('base.es').id,  # España
                        'type': 'contact',
                    }

                    # Teléfonos - Concatenar teléfono fijo y móvil si ambos existen (igual que proveedores/clientes)
                    if telefono and movil:
                        partner_vals['phone'] = f"{telefono} / {movil}"
                    elif telefono:
                        partner_vals['phone'] = telefono
                    elif movil:
                        partner_vals['phone'] = movil

                    # Obtener estado (provincia) desde código postal usando el mismo método que proveedores/clientes
                    if province_map and cod_postal and len(str(cod_postal)) >= 2:
                        try:
                            zip_str = str(cod_postal).strip()
                            if len(zip_str) >= 2:
                                provincia_codigo = zip_str[:2]
                                # Eliminar ceros a la izquierda para códigos como "01", "02", etc.
                                provincia_codigo = str(int(provincia_codigo))

                                if provincia_codigo in province_map:
                                    partner_vals['state_id'] = province_map[provincia_codigo]
                                    _logger.debug(f"✓ Provincia asignada desde CP {cod_postal}: código {provincia_codigo} -> state_id {province_map[provincia_codigo]}")
                                else:
                                    _logger.warning(f"✗ Código de provincia no encontrado: '{provincia_codigo}' extraído del CP: {cod_postal}")
                            else:
                                _logger.warning(f"✗ Código postal demasiado corto para extraer provincia: '{cod_postal}'")
                        except Exception as e:
                            _logger.warning(f"✗ Error al extraer provincia del código postal '{cod_postal}': {str(e)}")

                    # Crear o actualizar el contacto de trabajo
                    work_contact = None
                    if employee and employee.work_contact_id:
                        work_contact = employee.work_contact_id
                        if self.update_existing:
                            work_contact.write(partner_vals)
                    else:
                        # Buscar si existe un partner con este código
                        work_contact = self.env['res.partner'].search([
                            ('ref', '=', codigo)
                        ], limit=1) if codigo else None

                        if work_contact and self.update_existing:
                            work_contact.write(partner_vals)
                        elif not work_contact:
                            work_contact = self.env['res.partner'].create(partner_vals)

                    # Preparar valores para el empleado
                    employee_vals = {
                        'name': full_name,
                        'work_contact_id': work_contact.id,
                        'identification_id': nif if nif else False,
                        'ssnid': nss if nss else False,  # Número de Seguridad Social
                        'work_phone': telefono if telefono else False,
                        'mobile_phone': movil if movil else False,
                        'work_email': email if email else False,
                        # Dirección privada del empleado
                        'private_street': direccion if direccion else False,
                        'private_city': poblacion if poblacion else False,
                        'private_zip': cod_postal if cod_postal else False,
                        'private_country_id': self.env.ref('base.es').id,  # España
                    }

                    # Añadir provincia privada si existe
                    if province_map and cod_postal and len(str(cod_postal)) >= 2:
                        try:
                            zip_str = str(cod_postal).strip()
                            if len(zip_str) >= 2:
                                provincia_codigo = zip_str[:2]
                                provincia_codigo = str(int(provincia_codigo))
                                if provincia_codigo in province_map:
                                    employee_vals['private_state_id'] = province_map[provincia_codigo]
                        except Exception as e:
                            _logger.warning(f"✗ Error al extraer provincia para empleado {full_name}: {str(e)}")

                    # Crear o actualizar empleado
                    if employee:
                        if self.update_existing:
                            employee.write(employee_vals)
                            result['updated'] += 1
                            _logger.info(f"✓ Empleado actualizado: {full_name}")
                        else:
                            _logger.debug(f"◷ Empleado ya existe (no actualizar): {full_name}")
                    else:
                        employee = self.env['hr.employee'].create(employee_vals)
                        result['created'] += 1
                        _logger.info(f"✓ Empleado creado: {full_name}")

                    # Crear cuenta bancaria si existe
                    if cuenta_bancaria and cuenta_bancaria != '0':
                        self._create_employee_bank_account(employee, cuenta_bancaria)

                    # Crear o actualizar contrato si existe tipo de contrato
                    if tipo_contrato:
                        self._create_or_update_contract(employee, tipo_contrato, fecha_antiguedad)

                except Exception as e:
                    error_msg = f"Fila {row_idx}: {str(e)}"
                    result['errors'] += 1
                    result['error_list'].append(error_msg)
                    _logger.error(f"✗ {error_msg}")

        except Exception as e:
            error_msg = f"Error al procesar el archivo: {str(e)}"
            result['errors'] += 1
            result['error_list'].append(error_msg)
            _logger.error(f"✗ {error_msg}")

        return result

    def _create_employee_bank_account(self, employee, cuenta_bancaria):
        """Crea una cuenta bancaria para el empleado."""
        try:
            if not employee.work_contact_id:
                _logger.warning(f"El empleado {employee.name} no tiene contacto de trabajo asociado")
                return

            # Limpiar y formatear el número de cuenta
            acc_number = str(cuenta_bancaria).strip().replace(' ', '')

            if not acc_number or acc_number == '0':
                return

            # Buscar si ya existe una cuenta bancaria
            existing_bank = self.env['res.partner.bank'].search([
                ('partner_id', '=', employee.work_contact_id.id),
                ('acc_number', '=', acc_number),
            ], limit=1)

            if existing_bank:
                _logger.debug(f"Cuenta bancaria ya existe para {employee.name}")
                return

            # Crear nueva cuenta bancaria
            bank_vals = {
                'partner_id': employee.work_contact_id.id,
                'acc_number': acc_number,
            }

            # Si es un IBAN válido
            if acc_number and len(acc_number) > 2 and acc_number[:2].isalpha():
                bank_vals['acc_type'] = 'iban'

            self.env['res.partner.bank'].create(bank_vals)
            _logger.info(f"✓ Cuenta bancaria creada para empleado {employee.name}: {acc_number}")

        except Exception as e:
            _logger.warning(f"✗ Error creando cuenta bancaria para empleado {employee.name}: {str(e)}")

    def _create_or_update_contract(self, employee, tipo_contrato, fecha_antiguedad):
        """Crea o actualiza el contrato del empleado."""
        try:
            # Verificar si el módulo hr_contract está instalado
            if 'hr.contract' not in self.env:
                _logger.debug(f"Módulo hr_contract no instalado, saltando creación de contrato para {employee.name}")
                return

            # Buscar si ya existe un contrato activo para este empleado
            existing_contract = self.env['hr.contract'].search([
                ('employee_id', '=', employee.id),
                ('state', '=', 'open'),
            ], limit=1)

            contract_vals = {
                'name': f"Contrato - {employee.name}",
                'employee_id': employee.id,
            }

            # Añadir fecha de inicio si existe fecha de antigüedad
            if fecha_antiguedad:
                try:
                    # Intentar convertir la fecha
                    from datetime import datetime
                    if isinstance(fecha_antiguedad, str):
                        # Probar diferentes formatos de fecha
                        for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                            try:
                                date_obj = datetime.strptime(fecha_antiguedad, fmt)
                                contract_vals['date_start'] = date_obj.date()
                                break
                            except ValueError:
                                continue
                    elif hasattr(fecha_antiguedad, 'date'):
                        # Es un objeto datetime
                        contract_vals['date_start'] = fecha_antiguedad.date()
                    else:
                        contract_vals['date_start'] = fecha_antiguedad
                except Exception as e:
                    _logger.warning(f"No se pudo convertir la fecha de antigüedad para {employee.name}: {str(e)}")

            # Buscar o crear el tipo de contrato
            if tipo_contrato:
                if 'hr.contract.type' in self.env:
                    contract_type = self.env['hr.contract.type'].search([
                        ('name', '=ilike', tipo_contrato)
                    ], limit=1)

                    if not contract_type:
                        contract_type = self.env['hr.contract.type'].create({
                            'name': tipo_contrato
                        })
                        _logger.info(f"✓ Tipo de contrato creado: {tipo_contrato}")

                    contract_vals['contract_type_id'] = contract_type.id

            if existing_contract:
                if self.update_existing:
                    existing_contract.write(contract_vals)
                    _logger.info(f"✓ Contrato actualizado para {employee.name}")
            else:
                # Añadir el estado del contrato al crear uno nuevo
                contract_vals['state'] = 'open'
                self.env['hr.contract'].create(contract_vals)
                _logger.info(f"✓ Contrato creado para {employee.name}")

        except Exception as e:
            _logger.warning(f"✗ Error creando/actualizando contrato para {employee.name}: {str(e)}")

